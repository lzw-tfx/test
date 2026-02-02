"""
创建Excel导入模板
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def create_basic_info_template():
    """创建基本信息导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基本信息"
    
    headers = [
        '序号', '姓名*', '性别', '公民身份号码*', '民族', '政治面貌',
        '区划', '镇街', '连', '排', '班', '带训班长信息', '在营状态', '离营时间', '离营情况说明',
        '毕业或所在院校', '文化程度', '所学专业', '学业情况', '学习类型',
        '电话', '户籍地', '常住地', '家庭情况',
        '父母电话', '个人经历', '证明人', '证明人电话'
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # 添加示例数据
    example_data = [
        1, '张三', '男', '110101199001011234', '汉族', '团员',
        '朝阳区', '某某街道', '一连', '一排', '一班', '李班长', '在营', '', '正常',
        '某某大学', '本科', '计算机科学与技术', '在读', '全日制',
        '13800138000', '北京市朝阳区', '北京市朝阳区某某街道', '父母健在，家庭和睦',
        '13900139000', '2018-2022 某某大学计算机专业', '李老师', '13900139001'
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    os.makedirs('示例数据', exist_ok=True)
    wb.save('示例数据/基本信息导入模板.xlsx')
    print('✓ 基本信息导入模板创建成功')

def create_abnormal_stat_template():
    """创建异常情况统计模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "异常情况统计"
    
    headers = ['身份证号', '异常类型', '描述', '记录日期', '处理人', '状态']
    
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    example_data = [
        '110101199001011234', '身体不适', '训练中出现头晕', '2026-01-15', '李医生', '已处理'
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    wb.save('示例数据/异常情况统计模板.xlsx')
    print('✓ 异常情况统计模板创建成功')

def create_health_screening_template():
    """创建健康筛查模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "健康筛查"
    
    headers = ['身份证号', '筛查类型', '结果', '筛查日期', '后续跟进']
    
    header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    example_data = [
        '110101199001011234', '心理测评', '正常', '2026-01-10', '无需跟进'
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    wb.save('示例数据/健康筛查模板.xlsx')
    print('✓ 健康筛查模板创建成功')

def create_medical_screening_template():
    """创建病史筛查模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "病史筛查"
    
    headers = ['序号', '姓名*', '性别*', '公民身份号码*', '筛查情况*', '筛查日期*', '身体状况*', '精神状况*']
    
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    example_data = [
        1, '张三', '男', '110101199001011234', '无重大疾病史，无家族遗传病史', '2026-01-15', '正常', '正常'
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 调整列宽
    column_widths = [8, 12, 8, 20, 40, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    wb.save('示例数据/病史筛查模板.xlsx')
    print('✓ 病史筛查模板创建成功')

if __name__ == '__main__':
    print('正在创建Excel导入模板...\n')
    create_basic_info_template()
    create_abnormal_stat_template()
    create_health_screening_template()
    create_medical_screening_template()
    print('\n所有模板创建完成！文件位于 示例数据/ 目录')
